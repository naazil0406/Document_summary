import os
import sys
import tempfile
from openpyxl import Workbook

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import parse_and_chunk
from services.chunking import DocumentChunker
from services.excel_parser import ExcelParser


def test_excel_parser_extracts_messy_sheet_into_readable_text():
    with tempfile.TemporaryDirectory() as tmpdir:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Messy Data"
        sheet.append(["Region", "", "Sales"])
        sheet.append(["North", None, 1200])
        sheet.append([None, "Q1", 450])
        sheet.append(["South", "", 980])
        sheet.append(["", "Notes", "Delayed shipment"])

        path = os.path.join(tmpdir, "sample.xlsx")
        workbook.save(path)
        workbook.close()

        parser = ExcelParser(excel_folder=tmpdir)
        pages = parser.extract_pages(path)

        assert len(pages) == 1
        assert "Messy Data" in pages[0].page_label
        assert "Region" in pages[0].text
        assert "Sales" in pages[0].text
        assert "Delayed shipment" in pages[0].text
        assert "\n\n" in pages[0].text


def test_document_chunker_does_not_over_chunk_excel_sheet():
    with tempfile.TemporaryDirectory() as tmpdir:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Story Sheet"
        sheet.append(["Title", "Content"])
        sheet.append(["Intro", "Welcome to the onboarding story."])
        sheet.append(["Step 1", "Watch the first video carefully."])
        sheet.append(["Step 2", "Answer the reflection questions."])
        sheet.append(["Step 3", "Complete the quiz and submit."])

        path = os.path.join(tmpdir, "sample.xlsx")
        workbook.save(path)
        workbook.close()

        parser = ExcelParser(excel_folder=tmpdir)
        pages = parser.extract_pages(path)
        chunker = DocumentChunker(heading_max_length=80, min_paragraph_length=20)
        chunks = chunker.chunk_pages(pages)

        assert len(chunks) <= 3


def test_parse_and_chunk_can_skip_semantic_chunking_for_excel():
    with tempfile.TemporaryDirectory() as tmpdir:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Quick Sheet"
        sheet.append(["Item", "Value"])
        sheet.append(["A", 1])
        sheet.append(["B", 2])

        path = os.path.join(tmpdir, "sample.xlsx")
        workbook.save(path)
        workbook.close()

        class DummyEmbeddingService:
            @property
            def langchain_embeddings(self):
                return None

        chunks = parse_and_chunk(path, DummyEmbeddingService(), parser=ExcelParser(excel_folder=tmpdir), use_semantic_chunking=False)

        assert chunks is not None
        assert len(chunks) >= 1
        assert any("Item" in chunk.text for chunk in chunks)


class DummyEmbeddingService:
    @property
    def langchain_embeddings(self):
        return None


def test_excel_parser_processes_multiple_sheets_and_tables():
    with tempfile.TemporaryDirectory() as tmpdir:
        workbook = Workbook()
        sheet1 = workbook.active
        sheet1.title = "Employees"
        sheet1.append(["Name", "Department", "Salary"])
        sheet1.append(["John", "HR", 50000])
        sheet1.append(["Mary", "Sales", 60000])

        sheet1.append([])
        sheet1.append(["Name", "Location"])
        sheet1.append(["John", "Chennai"])

        sheet2 = workbook.create_sheet("Projects")
        sheet2.append(["Project", "Owner"])
        sheet2.append(["Alpha", "Nina"])

        path = os.path.join(tmpdir, "sample.xlsx")
        workbook.save(path)
        workbook.close()

        parser = ExcelParser(excel_folder=tmpdir)
        pages = parser.extract_pages(path)

        assert len(pages) >= 3
        assert any(page.metadata.get("sheet_name") == "Employees" for page in pages)
        assert any(page.metadata.get("sheet_name") == "Projects" for page in pages)


def test_excel_chunker_groups_rows_and_repeats_headers():
    with tempfile.TemporaryDirectory() as tmpdir:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Metrics"
        sheet.append(["Region", "Sales", "Quarter"])
        for idx in range(30):
            sheet.append([f"Region {idx % 3}", 100 + idx, "Q1"])

        path = os.path.join(tmpdir, "sample.xlsx")
        workbook.save(path)
        workbook.close()

        parser = ExcelParser(excel_folder=tmpdir)
        pages = parser.extract_pages(path)
        assert len(pages) >= 2

        chunks = parse_and_chunk(path, DummyEmbeddingService(), parser=parser, use_semantic_chunking=False)
        assert len(chunks) >= 2
        assert all("Columns:" in chunk.text for chunk in chunks)
        assert any(chunk.metadata.get("row_end", 0) >= 25 for chunk in chunks)


def test_large_multisheet_workbook_is_chunked_with_complete_metadata():
    with tempfile.TemporaryDirectory() as tmpdir:
        workbook = Workbook(write_only=True)
        for sheet_index in range(5):
            sheet = workbook.create_sheet(f"Department {sheet_index + 1}")
            sheet.append(["Employee", None, "Department", "Salary"])
            for row_index in range(600):
                sheet.append(
                    [
                        f"Employee {row_index + 1}",
                        None,
                        f"Department {sheet_index + 1}",
                        50000 + row_index,
                    ]
                )

        path = os.path.join(tmpdir, "enterprise.xlsx")
        workbook.save(path)

        pages = ExcelParser(excel_folder=tmpdir).extract_pages(path)

        assert len(pages) == 5 * 24
        assert {page.metadata["sheet_name"] for page in pages} == {
            f"Department {index}" for index in range(1, 6)
        }
        assert all(page.metadata["document_name"] == "enterprise.xlsx" for page in pages)
        assert all(page.metadata["table_name"] == "Table 1" for page in pages)
        assert all(1 <= page.metadata["row_start"] <= page.metadata["row_end"] for page in pages)
        assert all("Columns: Employee | Department | Salary" in page.text for page in pages)


def test_table_title_is_preserved_as_detected_table_name():
    with tempfile.TemporaryDirectory() as tmpdir:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Finance"
        sheet.append(["FY26 Operating Plan"])
        sheet.append(["Cost Center", "Owner", "Budget"])
        sheet.append(["Engineering", "Anita", 1250000])

        path = os.path.join(tmpdir, "plan.xlsx")
        workbook.save(path)
        workbook.close()

        pages = ExcelParser(excel_folder=tmpdir).extract_pages(path)

        assert len(pages) == 1
        assert pages[0].metadata["table_name"] == "FY26 Operating Plan"
        assert pages[0].metadata["row_start"] == 3
        assert "Cost Center: Engineering" in pages[0].text


def test_side_by_side_tables_are_detected_separately():
    with tempfile.TemporaryDirectory() as tmpdir:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Dashboard"
        sheet.append(["Employee", "Role", None, None, "Project", "Status"])
        sheet.append(["Ravi", "Engineer", None, None, "Apollo", "Active"])
        sheet.append(["Mina", "Manager", None, None, "Gemini", "Planned"])

        path = os.path.join(tmpdir, "dashboard.xlsx")
        workbook.save(path)
        workbook.close()

        pages = ExcelParser(excel_folder=tmpdir).extract_pages(path)

        assert len(pages) == 2
        assert {page.metadata["table_name"] for page in pages} == {
            "Table 1",
            "Table 2",
        }
        assert "Employee: Ravi" in pages[0].text
        assert "Project: Apollo" in pages[1].text
