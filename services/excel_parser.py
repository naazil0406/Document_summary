"""Structured, memory-conscious parsing for Excel workbooks and CSV files."""

from __future__ import annotations

import csv
import logging
import os
import re
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from typing import Iterable, Iterator, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

try:
    import xlrd  # type: ignore
except ImportError:  # pragma: no cover - optional legacy Excel dependency
    xlrd = None

from services.pdf_parser import PageContent

logger = logging.getLogger(__name__)

EXCEL_EXTENSIONS = (".xlsx", ".xlsm", ".xls")
CSV_EXTENSIONS = (".csv",)
TABULAR_EXTENSIONS = EXCEL_EXTENSIONS + CSV_EXTENSIONS
DEFAULT_ROWS_PER_CHUNK = 25
DEFAULT_MAX_CHUNK_CHARS = 6000
_NUMBER_RE = re.compile(r"^[-+]?(?:\d+(?:[,.]\d+)*|\.\d+)(?:[%$])?$")
_HEADER_LABELS = {
    "audio",
    "audio asset",
    "audio assets",
    "budget",
    "code",
    "content",
    "cindy s feedback",
    "department",
    "employee",
    "english",
    "english text",
    "feedback",
    "id",
    "illustrations",
    "ilustrations",
    "icons",
    "location",
    "module",
    "module name",
    "module no",
    "module number",
    "module type",
    "name",
    "number",
    "owner",
    "programming",
    "project",
    "prototype board no",
    "prototype board number",
    "quarter",
    "role",
    "salary",
    "sb no",
    "sb type",
    "status",
    "storyboard",
    "storyboard no",
    "storyboard number",
    "text",
    "title",
    "translation",
    "translation text",
    "type",
    "unit",
    "unit name",
    "unit no",
    "unit number",
    "ux",
    "value",
    "video",
    "video asset",
    "video assets",
}


@dataclass(frozen=True)
class _TableBlock:
    """A vertically contiguous group of non-empty worksheet rows."""

    start_row: int
    end_row: int
    rows: Tuple[Tuple[int, Tuple[object, ...]], ...]


def _cell_display(value: object) -> str:
    """Render common spreadsheet values predictably for embedding text."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_blank_row(row: Sequence[object]) -> bool:
    return not any(_cell_display(value) for value in row)


def _iter_table_blocks(
    rows: Iterable[Tuple[int, Sequence[object]]],
) -> Iterator[_TableBlock]:
    """Yield logical tables without mistaking layout spacers for boundaries.

    Spreadsheet authors frequently place blank rows between records. A new
    table begins only when the first populated row after a gap looks like an
    explicit column-header row.
    """
    current: List[Tuple[int, Tuple[object, ...]]] = []
    after_gap = False
    for row_number, values in rows:
        row = tuple(values)
        if _is_blank_row(row):
            after_gap = bool(current)
            continue
        if current and after_gap and _is_explicit_header_row(row):
            yield _TableBlock(current[0][0], current[-1][0], tuple(current))
            current = []
        current.append((row_number, row))
        after_gap = False
    if current:
        yield _TableBlock(current[0][0], current[-1][0], tuple(current))


def _is_explicit_header_row(row: Sequence[object]) -> bool:
    values = [_cell_display(value) for value in row if _cell_display(value)]
    if not values:
        return False
    label_cells = 0
    for value in values:
        normalized = " ".join(re.findall(r"[a-z]+", value.casefold()))
        if normalized in _HEADER_LABELS:
            label_cells += 1
    return label_cells >= 2 and label_cells / len(values) >= 0.4


def _remove_empty_columns(
    rows: Sequence[Tuple[int, Sequence[object]]],
) -> List[Tuple[int, List[str]]]:
    """Drop only columns that are empty throughout this table."""
    if not rows:
        return []
    width = max(len(values) for _, values in rows)
    keep = [
        column
        for column in range(width)
        if any(
            column < len(values) and bool(_cell_display(values[column]))
            for _, values in rows
        )
    ]
    return [
        (
            row_number,
            [
                _cell_display(values[column]) if column < len(values) else ""
                for column in keep
            ],
        )
        for row_number, values in rows
    ]


def _split_horizontal_tables(
    rows: Sequence[Tuple[int, Sequence[object]]],
) -> List[List[Tuple[int, Sequence[object]]]]:
    """Detect side-by-side tables separated by at least two empty columns.

    A single empty spacer column is common inside one visual table and is
    therefore removed later instead of being treated as a hard boundary.
    """
    if not rows:
        return []
    width = max(len(values) for _, values in rows)
    occupied = [
        any(
            column < len(values) and bool(_cell_display(values[column]))
            for _, values in rows
        )
        for column in range(width)
    ]
    populated = [index for index, has_value in enumerate(occupied) if has_value]
    if not populated:
        return []

    regions: List[Tuple[int, int]] = []
    region_start = populated[0]
    previous = populated[0]
    for column in populated[1:]:
        if column - previous >= 3:
            regions.append((region_start, previous))
            region_start = column
        previous = column
    regions.append((region_start, previous))

    return [
        [
            (row_number, tuple(values[start : end + 1]))
            for row_number, values in rows
        ]
        for start, end in regions
    ]


def _looks_like_header(
    candidate: Sequence[str],
    following_rows: Sequence[Sequence[str]],
) -> bool:
    populated = [value for value in candidate if value]
    if not populated:
        return False
    text_cells = sum(bool(re.search(r"[^\W\d_]", value, re.UNICODE)) for value in populated)
    numeric_cells = sum(bool(_NUMBER_RE.fullmatch(value)) for value in populated)
    if text_cells == 0 or numeric_cells > text_cells:
        return False

    # A title row normally has one populated cell followed by a wider row.
    next_width = max(
        (sum(bool(value) for value in row) for row in following_rows[:3]),
        default=0,
    )
    if len(populated) == 1 and next_width > 1:
        return False
    return True


def _unique_headers(values: Sequence[str]) -> List[str]:
    """Fill blank headers and disambiguate duplicates without losing order."""
    headers: List[str] = []
    counts: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        base = value or f"Column {index}"
        counts[base] = counts.get(base, 0) + 1
        headers.append(base if counts[base] == 1 else f"{base} ({counts[base]})")
    return headers


def _infer_english_text_header(
    headers: List[str],
    data_rows: Sequence[Tuple[int, List[str]]],
) -> List[str]:
    """Repair a missing English Text label in storyboard-style worksheets."""
    if any(header.casefold() == "english text" for header in headers):
        return headers

    normalized = {header.casefold() for header in headers}
    storyboard_schema = (
        any("module" in header for header in normalized)
        and any("storyboard" in header or header.startswith("sb ") for header in normalized)
    )
    if not storyboard_schema or not data_rows:
        return headers

    scores: List[Tuple[int, int, int]] = []
    for column in range(len(headers)):
        values = [
            row[column]
            for _, row in data_rows
            if column < len(row) and row[column]
        ]
        total_chars = sum(len(value) for value in values)
        long_cells = sum(len(value) >= 40 for value in values)
        scores.append((long_cells, total_chars, column))

    long_cells, total_chars, column = max(scores, default=(0, 0, -1))
    if column >= 0 and long_cells >= 3 and total_chars >= 300:
        repaired = list(headers)
        repaired[column] = "English Text"
        logger.info(
            "Inferred missing 'English Text' header for column %d (previous label: %r).",
            column + 1,
            headers[column],
        )
        return _unique_headers(repaired)
    return headers


def _fill_context_columns(
    headers: Sequence[str],
    rows: Sequence[Tuple[int, List[str]]],
) -> List[Tuple[int, List[str]]]:
    """Carry record identity into continuation rows with narration-only cells."""
    context_columns = [
        index
        for index, header in enumerate(headers)
        if any(
            term in header.casefold()
            for term in (
                "unit no",
                "unit name",
                "module no",
                "module name",
                "module type",
                "storyboard no",
                "prototype board no",
                "sb no",
                "sb type",
            )
        )
    ]
    previous: dict[int, str] = {}
    enriched: List[Tuple[int, List[str]]] = []
    for row_number, values in rows:
        updated = list(values)
        for column in context_columns:
            if column >= len(updated):
                continue
            if updated[column]:
                previous[column] = updated[column]
            elif column in previous:
                updated[column] = previous[column]
        enriched.append((row_number, updated))
    return enriched


def _table_layout(
    rows: Sequence[Tuple[int, List[str]]],
    table_index: int,
) -> Tuple[str, List[str], List[Tuple[int, List[str]]]]:
    """Find an optional title, a header row, and the data rows."""
    if not rows:
        return "", [], []

    header_index: Optional[int] = None
    search_limit = min(len(rows), 10)
    for index in range(search_limit):
        if _looks_like_header(rows[index][1], [row for _, row in rows[index + 1 : index + 4]]):
            header_index = index
            break

    if header_index is None:
        width = max(len(values) for _, values in rows)
        return f"Table {table_index}", _unique_headers([""] * width), list(rows)

    title = ""
    if header_index > 0:
        title_parts = [value for value in rows[header_index - 1][1] if value]
        candidate = " ".join(title_parts).strip()
        if candidate and len(candidate) <= 120:
            title = candidate

    data_rows = list(rows[header_index + 1 :])
    headers = _unique_headers(rows[header_index][1])
    headers = _infer_english_text_header(headers, data_rows)
    data_rows = _fill_context_columns(headers, data_rows)
    return title or f"Table {table_index}", headers, data_rows


def _build_structured_chunks(
    *,
    filename: str,
    sheet_name: str,
    sheet_index: int,
    table_rows: Sequence[Tuple[int, Sequence[object]]],
    table_index: int,
    rows_per_chunk: int,
    max_chunk_chars: int,
) -> List[PageContent]:
    cleaned = _remove_empty_columns(table_rows)
    table_name, headers, data_rows = _table_layout(cleaned, table_index)
    data_rows = [(number, values) for number, values in data_rows if any(values)]
    if not headers or not data_rows:
        logger.debug(
            "Skipping header-only/empty table %d on sheet '%s'.",
            table_index,
            sheet_name,
        )
        return []

    batches: List[List[Tuple[int, List[str]]]] = []
    current: List[Tuple[int, List[str]]] = []
    current_chars = 0
    for row in data_rows:
        row_chars = sum(
            len(header) + len(value) + 3
            for header, value in zip(headers, row[1])
            if value
        )
        if current and (
            len(current) >= rows_per_chunk
            or current_chars + row_chars > max_chunk_chars
        ):
            batches.append(current)
            current = []
            current_chars = 0
        current.append(row)
        current_chars += row_chars
    if current:
        batches.append(current)

    pages: List[PageContent] = []
    for batch in batches:
        row_start, row_end = batch[0][0], batch[-1][0]
        lines = [
            f"Sheet: {sheet_name}",
            f"Table: {table_name}",
            "Columns: " + " | ".join(headers),
            "",
        ]
        for row_number, values in batch:
            lines.append(f"Row {row_number}")
            pairs = [
                f"{header}: {value}"
                for header, value in zip(headers, values)
                if value
            ]
            lines.extend(pairs or ["Values: none"])

        metadata = {
            "document_name": filename,
            "sheet_name": sheet_name,
            "table_name": table_name,
            "row_start": row_start,
            "row_end": row_end,
            "toc_section": sheet_name,
        }
        pages.append(
            PageContent(
                filename=filename,
                page_number=sheet_index,
                page_label=f"Sheet: {sheet_name} | Table: {table_name}",
                text="\n".join(lines),
                metadata=metadata,
            )
        )
    return pages


class ExcelParser:
    """Emit Excel/CSV row groups through the pipeline's PageContent contract."""

    def __init__(
        self,
        excel_folder: str,
        llm_service: Optional[object] = None,
        rows_per_chunk: int = DEFAULT_ROWS_PER_CHUNK,
        max_chunk_chars: int = DEFAULT_MAX_CHUNK_CHARS,
    ):
        if not 20 <= rows_per_chunk <= 30:
            raise ValueError("rows_per_chunk must be between 20 and 30")
        self.excel_folder = excel_folder
        # Kept for backward compatibility with callers from older releases.
        self.llm_service = llm_service
        self.rows_per_chunk = rows_per_chunk
        self.max_chunk_chars = max(1000, max_chunk_chars)

    def _list_excel_files(self) -> List[str]:
        if not os.path.isdir(self.excel_folder):
            return []
        return [
            os.path.join(self.excel_folder, name)
            for name in sorted(os.listdir(self.excel_folder))
            if name.lower().endswith(TABULAR_EXTENSIONS)
        ]

    def _chunks_from_rows(
        self,
        rows: Iterable[Tuple[int, Sequence[object]]],
        filename: str,
        sheet_name: str,
        sheet_index: int,
    ) -> List[PageContent]:
        pages: List[PageContent] = []
        table_index = 0
        for block in _iter_table_blocks(rows):
            for table_rows in _split_horizontal_tables(block.rows):
                table_index += 1
                pages.extend(
                    _build_structured_chunks(
                        filename=filename,
                        sheet_name=sheet_name,
                        sheet_index=sheet_index,
                        table_rows=table_rows,
                        table_index=table_index,
                        rows_per_chunk=self.rows_per_chunk,
                        max_chunk_chars=self.max_chunk_chars,
                    )
                )
        return pages

    def _extract_workbook_pages(self, file_path: str, filename: str) -> List[PageContent]:
        try:
            workbook = load_workbook(
                file_path,
                data_only=True,
                read_only=True,
                keep_links=False,
            )
        except Exception:
            logger.exception("Could not open Excel workbook '%s'.", filename)
            return []

        pages: List[PageContent] = []
        try:
            for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
                try:
                    rows = (
                        (row_number, values)
                        for row_number, values in enumerate(
                            sheet.iter_rows(values_only=True),
                            start=1,
                        )
                    )
                    sheet_pages = self._chunks_from_rows(
                        rows, filename, sheet.title, sheet_index
                    )
                    pages.extend(sheet_pages)
                    logger.info(
                        "Parsed sheet '%s' from '%s' into %d chunk(s).",
                        sheet.title,
                        filename,
                        len(sheet_pages),
                    )
                except Exception:
                    logger.exception(
                        "Failed to parse sheet '%s' in '%s'.",
                        sheet.title,
                        filename,
                    )
        finally:
            workbook.close()
        return pages

    @staticmethod
    def _xls_value(workbook, cell) -> object:
        if xlrd is not None and cell.ctype == xlrd.XL_CELL_DATE:
            try:
                return xlrd.xldate_as_datetime(cell.value, workbook.datemode)
            except (TypeError, ValueError):
                return cell.value
        if xlrd is not None and cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return bool(cell.value)
        return cell.value

    def _extract_workbook_pages_xls(
        self, file_path: str, filename: str
    ) -> List[PageContent]:
        if xlrd is None:
            raise RuntimeError(
                "Legacy .xls support requires xlrd. Install dependencies from requirements.txt."
            )
        try:
            workbook = xlrd.open_workbook(file_path, on_demand=True)
        except Exception:
            logger.exception("Could not open legacy Excel workbook '%s'.", filename)
            return []

        pages: List[PageContent] = []
        try:
            for sheet_index, sheet_name in enumerate(workbook.sheet_names(), start=1):
                sheet = workbook.sheet_by_name(sheet_name)
                rows = (
                    (
                        row_index + 1,
                        [
                            self._xls_value(workbook, sheet.cell(row_index, column))
                            for column in range(sheet.ncols)
                        ],
                    )
                    for row_index in range(sheet.nrows)
                )
                sheet_pages = self._chunks_from_rows(
                    rows, filename, sheet_name, sheet_index
                )
                pages.extend(sheet_pages)
                workbook.unload_sheet(sheet_name)
        finally:
            workbook.release_resources()
        return pages

    def _extract_csv_pages(self, file_path: str, filename: str) -> List[PageContent]:
        try:
            with open(file_path, "r", encoding="utf-8-sig", newline="") as handle:
                sample = handle.read(4096)
                handle.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
                except csv.Error:
                    dialect = csv.excel
                rows = (
                    (row_number, row)
                    for row_number, row in enumerate(csv.reader(handle, dialect), start=1)
                )
                return self._chunks_from_rows(rows, filename, "CSV Data", 1)
        except (OSError, UnicodeError, csv.Error):
            logger.exception("Could not parse CSV file '%s'.", filename)
            return []

    def extract_pages(self, file_path: str) -> List[PageContent]:
        filename = os.path.basename(file_path)
        extension = os.path.splitext(filename)[1].lower()
        if extension not in TABULAR_EXTENSIONS:
            raise ValueError(f"Unsupported tabular file type: {extension or '<none>'}")
        if extension in CSV_EXTENSIONS:
            return self._extract_csv_pages(file_path, filename)
        if extension == ".xls":
            return self._extract_workbook_pages_xls(file_path, filename)
        return self._extract_workbook_pages(file_path, filename)

    def extract_all(self) -> List[PageContent]:
        pages: List[PageContent] = []
        files = self._list_excel_files()
        logger.info("Found %d tabular file(s) in '%s'.", len(files), self.excel_folder)
        for path in files:
            try:
                extracted = self.extract_pages(path)
                pages.extend(extracted)
                logger.info(
                    "Extracted %d structured chunk(s) from '%s'.",
                    len(extracted),
                    os.path.basename(path),
                )
            except Exception:
                logger.exception("Failed to process tabular file '%s'.", path)
        return pages
